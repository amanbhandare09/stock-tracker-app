# import openpyxl
# import pandas as pd
# import os
# import re
# from datetime import datetime
# import subprocess

# def clean_excel_sheet(sheet):
#     data = []
#     headers_found = False
#     content_start = False
#     header_row = []
#     first_blank = False
#     header_pattern = r'Name of (the )?Instrument(s)?'
#     equity_related_pattern = r'.*\b[eE]quity\s*(&|and)\s*[eE]quity\s*[rR]elated\b.*'
#     listed_pattern = r'.*\b[Ll]isted\s*/\s*[aA]waiting\s*[lL]isting\s*on\s*(the\s*)?[sS]tock\s*[eE]xchange.*'

#     desired_columns_patterns = {
#         "instrument": r'Name of (the )?Instrument(s)?',
#         "isin": r'ISIN',
#         "quantity": r'Quantity',
#         "holding_percentage": r'%\s*(to|of)?\s*Net\s*Asset(s)?|Rounded\s*%\s*(to|of)?\s*Net\s*Asset(s)?|%\s*(to|of)?\s*NAV|%\s*of\s*AUM|%\s*to\s*AUM|% age to NAV',
#     }

#     for row in sheet.iter_rows(values_only=True):
#         row_str = str(row)

#         if not headers_found:
#             if re.search(header_pattern, row_str, re.IGNORECASE):
#                 headers_found = True
#                 header_row = row
#                 continue
#             else:
#                 continue

#         if headers_found:
#             if isinstance(str(row), str) and re.search(equity_related_pattern, str(row), re.IGNORECASE):
#                 content_start = True
#                 continue

#             if content_start and isinstance(str(row), str) and re.search(listed_pattern, str(row), re.IGNORECASE):
#                 continue

#             if content_start and ("Arbitrage & Special Situations" or "Equity"  ) in str(row):
#                 continue

#             if content_start:
#                 if not any(row):
#                     if not first_blank:
#                         first_blank = True
#                         continue
#                     else:
#                         break
#                 data.append(row)

#     header_row = [str(col) if col is not None else "" for col in header_row]

#     df = pd.DataFrame(data, columns=header_row)

#     df = df[~df.apply(lambda row: row.astype(str).str.contains('NIL').any(), axis=1)]

#     column_mapping = {}
#     for col in df.columns:
#         col_str = str(col) if col is not None else ""
#         for key, pattern in desired_columns_patterns.items():
#             if re.search(pattern, col_str, re.IGNORECASE):
#                 column_mapping[col] = key
#                 break

#     df.rename(columns=column_mapping, inplace=True)

#     df = df[[key for key in desired_columns_patterns.keys() if key in df.columns]]

#     # Search for "Sub Total" or similar keywords across all columns
#     sub_total_index = None
#     for col in df.columns:
#         match_index = df[df[col].astype(str).str.contains(r'\b(?:Sub\s*Total|Subtotal|Total)\b', na=False, case=False)].index
#         if not match_index.empty:
#             sub_total_index = match_index[0]
#             break

#     if sub_total_index is not None:
#         df = df.iloc[:sub_total_index]

#     # Drop rows where both 'isin' and 'quantity' columns are empty
#     if 'isin' in df.columns and 'quantity' in df.columns:
#         df = df.dropna(subset=['isin', 'quantity'], how='all')

#     # Drop rows where 'quantity' column contains non-numeric values
#     if 'quantity' in df.columns:
#         df = df[pd.to_numeric(df['quantity'], errors='coerce').notnull()]

#     return df

# def sheet_contains_equity_related(sheet):
#     equity_related_pattern = r'.*\b[eE]quity\s*(&|and)\s*[eE]quity\s*[rR]elated\b.*'
#     listed_pattern = r'.*\b[Ll]isted\s*/\s*[aA]waiting\s*[lL]isting\s*on\s*(the\s*)?[sS]tock\s*[eE]xchange.*'
#     found_equity_related = False

#     for row in sheet.iter_rows(values_only=True):
#         if found_equity_related:
#             if not any(row):
#                 return False
#             if any(cell and isinstance(cell, str) and re.search(listed_pattern, cell, re.IGNORECASE) for cell in row):
#                 return True
#             return False

#         if any(cell and isinstance(cell, str) and re.search(equity_related_pattern, cell, re.IGNORECASE) for cell in row):
#             found_equity_related = True
    
#     return False

# def extract_scheme_name(scheme_name, amc_name):
#     if not isinstance(scheme_name, str):
#         scheme_name = str(scheme_name)
#         # For other AMCs, use the general extraction
#         match = re.match(r'^[\w\s-]+(?=\W|$)', scheme_name)
#         if match:
#             return match.group(0).strip()
#     return scheme_name

# def process_excel_file(file_path, schemeRow, schemeCol, amc_name ):
#     wb = openpyxl.load_workbook(file_path, data_only=True)
#     all_data = []

#     for sheet_name in wb.sheetnames:
#         sheet = wb[sheet_name]
        
#         print(f"Processing {sheet_name}.")

#         scheme_name = sheet.cell(row=schemeRow, column=schemeCol).value

#         if scheme_name:
#             scheme_name = extract_scheme_name(scheme_name,amc_name)
#         print(f'Scheme name: {scheme_name}, AMC name: {amc_name}')
        
#         if sheet_contains_equity_related(sheet):
#             cleaned_data = clean_excel_sheet(sheet )
#             if cleaned_data is not None and not cleaned_data.empty:
#                 if "% Net Asset" in cleaned_data.columns:
#                     cleaned_data = cleaned_data[cleaned_data["% Net Asset"] != "$0.00%"]
                
#                 if not cleaned_data.empty:
#                     cleaned_data['amc_short_name'] = amc_name
#                     cleaned_data['mf_short_name'] = sheet_name
#                     cleaned_data['fund_name'] = scheme_name
#                     cleaned_data['fund_type'] = 'equity'
#                     cleaned_data['month_year'] = datetime(2024, 6, 1).strftime('%b-%Y')
#                     all_data.append(cleaned_data.reset_index(drop=True))
#         else:
#             print(f"Skipping sheet {sheet_name} as it does not contain 'Equity & Equity related' or a similar suffix.")

#     if all_data:
#         combined_df = pd.concat(all_data, ignore_index=True)
#         return combined_df
#     else:
#         return pd.DataFrame()

# def process_amc_directory(directory_path, output_file_path):
#     all_amc_data = []
#     amc = {
#         "SUNDARAM":(2,1),
#     }

#     for amc_name, (row, col ) in amc.items():
#         sub_dir = os.path.join(directory_path, amc_name)
#         if not os.path.isdir(sub_dir):
#             print(f"Directory {sub_dir} not found, skipping.")
#             continue

#         all_data = []
#         for file in os.listdir(sub_dir):
#             if file.endswith('.xlsx'):
#                 file_path = os.path.join(sub_dir, file)
#                 print(f"Processing file: {file} in {sub_dir}")
#                 amc_data = process_excel_file(file_path, row, col, amc_name )
#                 print(f"Processed data shape: {amc_data.shape if amc_data is not None else 'No data'}")
#                 if amc_data is not None and not amc_data.empty:
#                     all_data.append(amc_data)

#         if all_data:
#             combined_df = pd.concat(all_data, ignore_index=True)
#             output_file = os.path.join(output_file_path, f"{amc_name}.csv")
#             combined_df.to_csv(output_file, index=False)
#             print(f"Combined data saved to {output_file}")
#             all_amc_data.append(combined_df)
#         else:
#             print(f"No valid data found for AMC {amc_name}.")

#     if all_amc_data:
#         combined_all_amc_df = pd.concat(all_amc_data, ignore_index=True)
#         combined_all_amc_df.to_csv(output_file_path, index=False)
#         print(f"Combined data saved to {output_file_path}")
#     else:
#         print("No data to save in the combined output file.")

# if __name__ == "__main__":
#     import sys
#     if len(sys.argv) != 3:
#         print("Usage: python script_name.py <input_directory_path> <output_file_path>")
#         sys.exit(1)

#     SOURCE_FOLDER = sys.argv[1]
#     OUTPUT_FILE = sys.argv[2]

#     process_amc_directory(SOURCE_FOLDER, OUTPUT_FILE)
 


import openpyxl
import pandas as pd
import os
import re
from datetime import datetime
import sys

# Define the function to clean the Excel sheet data
def clean_excel_sheet(sheet):
    data = []
    headers_found = False
    content_start = False
    header_row = []
    first_blank = False
    header_pattern = r'Name of (the )?Instrument(s)?'
    equity_related_pattern = r'.*\b[eE]quity\s*(&|and)\s*[eE]quity\s*[rR]elated\b.*'
    listed_pattern = r'.*\b[Ll]isted\s*/\s*[aA]waiting\s*[lL]isting\s*on\s*(the\s*)?[sS]tock\s*[eE]xchange.*'

    desired_columns_patterns = {
        "instrument": r'Name of (the )?Instrument(s)?',
        "isin": r'ISIN',
        "quantity": r'Quantity',
        "holding_percentage": r'%\s*(to|of)?\s*Net\s*Asset(s)?|Rounded\s*%\s*(to|of)?\s*Net\s*Asset(s)?|%\s*(to|of)?\s*NAV|%\s*of\s*AUM|%\s*to\s*AUM|% age to NAV',
    }

    for row in sheet.iter_rows(values_only=True):
        row_str = str(row)

        if not headers_found:
            if re.search(header_pattern, row_str, re.IGNORECASE):
                headers_found = True
                header_row = row
                continue
            else:
                continue

        if headers_found:
            if isinstance(str(row), str) and re.search(equity_related_pattern, str(row), re.IGNORECASE):
                content_start = True
                continue

            if content_start and isinstance(str(row), str) and re.search(listed_pattern, str(row), re.IGNORECASE):
                continue

            if content_start and ("Arbitrage & Special Situations" or "Equity") in str(row):
                continue

            if content_start:
                if not any(row):
                    if not first_blank:
                        first_blank = True
                        continue
                    else:
                        break
                data.append(row)

    header_row = [str(col) if col is not None else "" for col in header_row]

    df = pd.DataFrame(data, columns=header_row)

    df = df[~df.apply(lambda row: row.astype(str).str.contains('NIL').any(), axis=1)]

    column_mapping = {}
    for col in df.columns:
        col_str = str(col) if col is not None else ""
        for key, pattern in desired_columns_patterns.items():
            if re.search(pattern, col_str, re.IGNORECASE):
                column_mapping[col] = key
                break

    df.rename(columns=column_mapping, inplace=True)

    df = df[[key for key in desired_columns_patterns.keys() if key in df.columns]]

    # Search for "Sub Total" or similar keywords across all columns
    sub_total_index = None
    for col in df.columns:
        match_index = df[df[col].astype(str).str.contains(r'\b(?:Sub\s*Total|Subtotal|Total)\b', na=False, case=False)].index
        if not match_index.empty:
            sub_total_index = match_index[0]
            break

    if sub_total_index is not None:
        df = df.iloc[:sub_total_index]

    # Drop rows where both 'isin' and 'quantity' columns are empty
    if 'isin' in df.columns and 'quantity' in df.columns:
        df = df.dropna(subset=['isin', 'quantity'], how='all')

    # Drop rows where 'quantity' column contains non-numeric values
    if 'quantity' in df.columns:
        df = df[pd.to_numeric(df['quantity'], errors='coerce').notnull()]

    return df

# Check if the sheet contains 'Equity & Equity related' content
def sheet_contains_equity_related(sheet):
    equity_related_pattern = r'.*\b[eE]quity\s*(&|and)\s*[eE]quity\s*[rR]elated\b.*'
    listed_pattern = r'.*\b[Ll]isted\s*/\s*[aA]waiting\s*[lL]isting\s*on\s*(the\s*)?[sS]tock\s*[eE]xchange.*'
    found_equity_related = False

    for row in sheet.iter_rows(values_only=True):
        if found_equity_related:
            if not any(row):
                return False
            if any(cell and isinstance(cell, str) and re.search(listed_pattern, cell, re.IGNORECASE) for cell in row):
                return True
            return False

        if any(cell and isinstance(cell, str) and re.search(equity_related_pattern, cell, re.IGNORECASE) for cell in row):
            found_equity_related = True
    
    return False

# Extract the scheme name
def extract_scheme_name(scheme_name, amc_name):
    if not isinstance(scheme_name, str):
        scheme_name = str(scheme_name)
        # For other AMCs, use the general extraction
        match = re.match(r'^[\w\s-]+(?=\W|$)', scheme_name)
        if match:
            return match.group(0).strip()
    return scheme_name

# Process a single Excel file
def process_excel_file(file_path, schemeRow, schemeCol, amc_name ):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_data = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        print(f"Processing {sheet_name}.")

        scheme_name = sheet.cell(row=schemeRow, column=schemeCol).value

        if scheme_name:
            scheme_name = extract_scheme_name(scheme_name, amc_name)
        print(f'Scheme name: {scheme_name}, AMC name: {amc_name}')
        
        if sheet_contains_equity_related(sheet):
            cleaned_data = clean_excel_sheet(sheet)
            if cleaned_data is not None and not cleaned_data.empty:
                if "% Net Asset" in cleaned_data.columns:
                    cleaned_data = cleaned_data[cleaned_data["% Net Asset"] != "$0.00%"]
                if "holding_percentage" in cleaned_data.columns:
                  # Replace specific values in the 'holding_percentage' column
                    cleaned_data['holding_percentage'] = cleaned_data['holding_percentage'].replace({ '#': '0.00'})

                if "holding_percentage" in cleaned_data.columns:
                    cleaned_data["holding_percentage"] = (cleaned_data["holding_percentage"].astype(float) * 100).round(2)

                if not cleaned_data.empty:
                    cleaned_data['amc_short_name'] = amc_name
                    cleaned_data['mf_short_name'] = sheet_name
                    cleaned_data['fund_name'] = scheme_name
                    cleaned_data['fund_type'] = 'equity'
                    cleaned_data['month_year'] = (datetime.now() - pd.DateOffset(months=1)).strftime('%b-%Y')
                    all_data.append(cleaned_data.reset_index(drop=True))
        else:
            print(f"Skipping sheet {sheet_name} as it does not contain 'Equity & Equity related' or a similar suffix.")

    if all_data:
        combined_df = pd.concat(all_data, ignore_index=True)
        return combined_df
    else:
        return pd.DataFrame()

 

# Process the AMC directory
def process_amc_directory(directory_path, output_file_path):
    all_amc_data = []
    amc = {
        "SUNDARAM": (2, 1),
    }

    for amc_name, (row, col ) in amc.items():
        sub_dir = os.path.join(directory_path)  # Updated to directly access the directory
        if not os.path.isdir(sub_dir):
            print(f"Directory {sub_dir} not found, skipping.")
            continue

        all_data = []
        for file in os.listdir(sub_dir):
            if file.endswith('.xlsx'):
                file_path = os.path.join(sub_dir, file)
                print(f"Processing file: {file} in {sub_dir}")
                amc_data = process_excel_file(file_path, row, col, amc_name)
                print(f"Processed data shape: {amc_data.shape if amc_data is not None else 'No data'}")
                if amc_data is not None and not amc_data.empty:
                    all_data.append(amc_data)

        if all_data:
            combined_df = pd.concat(all_data, ignore_index=True)
            
            # Ensure the output directory exists
            output_dir = os.path.dirname(output_file_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                
            output_file = os.path.join(output_file_path)
            combined_df.to_csv(output_file, index=False)
            print(f"Combined data saved to {output_file}")
            all_amc_data.append(combined_df)
        else:
            print(f"No valid data found for AMC {amc_name}.")

    if all_amc_data:
        combined_all_amc_df = pd.concat(all_amc_data, ignore_index=True)
        combined_all_amc_df.to_csv(output_file_path, index=False)
        print(f"Combined data saved to {output_file_path}")
    else:
        print("No data to save in the combined output file.")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python script_name.py <directory_path> <output_file_path>")
        sys.exit(1)

    directory_path = sys.argv[1]
    output_file_path = sys.argv[2]
    
    process_amc_directory(directory_path, output_file_path)
