# For BOI
import openpyxl
import pandas as pd
import os
import re
import glob

def clean_excel_sheet(sheet):
    data = []
    headers_found = False
    content_start = False
    relevant_columns = ["Name of the Instrument", "ISIN", "Quantity", "% to Net Assets"]
    exclude_pattern = r'364 Days Tbill.*|Tbill.*|.*Tbill'

    for row in sheet.iter_rows(values_only=True):
        if not headers_found:
            if "Name of the Instrument" in str(row):
                headers_found = True
                header_row = row
                continue
            else:
                continue

        if headers_found:
            if "Equity & Equity related" in str(row):
                content_start = True
                continue

            if content_start and "(a) Listed / awaiting listing on Stock Exchanges" in str(row):
                continue

            if content_start:
                if not any(cell for cell in row):
                    break
                data.append(row)

    header_row = [str(col) if col is not None else "" for col in header_row]
    df = pd.DataFrame(data, columns=header_row)

    df = df[[col for col in relevant_columns if col in df.columns]]

    # df = df[df['Quantity'] >= 0]

    df = df[~df['Name of the Instrument'].str.contains(exclude_pattern, na=False, case=False)]

    df = df[~df.apply(lambda row: row.astype(str).str.contains('NIL').any(), axis=1)]

    sub_total_index = df[df['Name of the Instrument'].str.contains('Sub Total', na=False)].index
    if not sub_total_index.empty:
        sub_total_index = sub_total_index[0] + 1
        df = df.iloc[:sub_total_index]

    return df

def extract_scheme_name(scheme_name):
    # Regular expression to capture text up to the first special character
    match = re.match(r'^[^\W_]+(?: [^\W_]+)*(?=\W)', scheme_name)
    if match:
        return match.group(0)
    return scheme_name

def process_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    all_data = []

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Index':
            continue

        sheet = wb[sheet_name]

        print(f"Processing {sheet_name}.")
        cell_value = sheet.cell(row=7, column=2).value
        scheme_name = sheet.cell(row=1, column=2).value
        if scheme_name:
            scheme_name = extract_scheme_name(scheme_name)
        if cell_value == "Equity & Equity related":
            cleaned_data = clean_excel_sheet(sheet)
            if cleaned_data is not None and not cleaned_data.empty:
                cleaned_data['amc_name'] = 'BOI'
                cleaned_data['Short name'] = sheet_name
                cleaned_data['Scheme Name'] = scheme_name
                all_data.append(cleaned_data)
        else:
            print(f"Skipping sheet {sheet_name} as it does not contain 'Equity & Equity related' or a similar suffix.")

    if all_data:
        combined_df = pd.concat(all_data, ignore_index=True)
        return combined_df
    else:
        print(f"No relevant data found in any sheet of file {file_path}.")
        return None

def process_amc_directory(directory_path, output_file_path):
    all_amc_data = []

    for file in os.listdir(directory_path):
        if file.endswith('.xlsx'):
            file_path = os.path.join(directory_path, file)
            print(f"Processing file: {file}")
            amc_data = process_excel_file(file_path)
            if amc_data is not None:
                all_amc_data.append(amc_data)

    if not all_amc_data:
        print("No valid data found in any Excel file.")
        return

    # Combine data from all AMCs and Excel files
    combined_df = pd.concat(all_amc_data, ignore_index=True)
    combined_df.to_csv(output_file_path, index=False)
    print(f"Combined data saved to {output_file_path}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Usage: python script_name.py <input_directory_path> <output_file_path>")
        sys.exit(1)

    SOURCE_FOLDER = sys.argv[1]
    OUTPUT_FILE = sys.argv[2]

    process_amc_directory(SOURCE_FOLDER, OUTPUT_FILE)
