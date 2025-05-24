import openpyxl
import numpy as np
import pandas as pd
import os
import re
import subprocess
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

MONTHLY_DIRECTORY_PATH = os.getenv('MONTHLY_DIRECTORY_PATH')
MONTHLY_OUTPUT_PATH = os.getenv('MONTHLY_OUTPUT_PATH')

def extract_scheme_name(scheme_name, amc_name):
    if not isinstance(scheme_name, str):
        scheme_name = str(scheme_name)
        
    if amc_name in ["EDELWEISS", "KOTAK", "UNION"]:
        scheme_name = scheme_name.lower()
        match = re.search(r'of\s+(.*?)\s+as\s+on', scheme_name)
        if match:
            return match.group(1).strip()
        
    elif amc_name == "GROWW":
        if "Scheme -" in scheme_name:
            match = re.search(r'-\s*([^(]*)', scheme_name)
            if match:
                return match.group(1).strip()
        else:
            match = re.match(r'^[^(]*', scheme_name)
            if match:
                return match.group(0).strip()
            
    elif amc_name == 'ZERODHA':
        scheme_name = scheme_name.lower()
        match = re.search(r'of\s+(.*?)\s+for', scheme_name)
        if match:
            return match.group(1).strip()
        
    elif amc_name == "TAURUS":
        return ' '.join(scheme_name.split())
    
    else:
        match = re.match(r'^[\w\s-]+(?=\W|$)', scheme_name)
        if match:
            return match.group(0).strip()
    
    return scheme_name

def sheet_contains_equity_related(sheet):
    equity_related_pattern = r'^[eE]quity.*'
    listed_pattern = r'.*\b[Ll]isted\s*/\s*[aA]waiting\s*[lL]isting\s*on\s*(the\s*)?[sS]tock\s*[eE]xchange.*'
    found_equity_related = False
    first_blank = False

    for row in sheet.iter_rows(values_only=True):
        row_str = " ".join([str(cell) if cell else "" for cell in row])
        row_str = re.sub(r'\d+(\.\d+)?', '', row_str).strip()

        if found_equity_related:
            if not any(row):
                if not first_blank:
                    first_blank = True
                    continue
                else:
                    print("empty row found")
                    return False
            
            if row_str == 'Equity Shares':
                continue

            if re.search(listed_pattern, row_str, re.IGNORECASE):
                return True
            
            return False
        
        if re.search(equity_related_pattern, row_str, re.IGNORECASE):
            found_equity_related = True
            
    return False

def clean_excel_sheet(sheet):
    data = []
    headers_found = False
    content_start = False
    header_row = []
    first_blank = False

    header_pattern = r'(Company/Issuer/Instrument Name)|(Name of (the )?Instrument(s)?)'
    equity_related_pattern = r'^[A]\)\s*Equity\s*&\s*Equity\s*Related|^[eE]quity.*'
    listed_pattern = r'.*\b[Ll]isted\s*/\s*[aA]waiting\s*[lL]isting\s*on\s*(the\s*)?[sS]tock\s*[eE]xchange.*'
    stop_pattern = r'TOTAL:\(a\)\s*[Ll]isted\s*/\s*[Aa]waiting\s*[Ll]isting\s*on\s*[Ss]tock\s*[Ee]xchange'

    desired_columns_patterns = {
        "instrument": r'(Company/Issuer/Instrument Name)|(Name of (the )?Instrument(s)?)',
        "isin": r'ISIN|ISIN Code',
        "quantity": r'Quantity',
        "holding_percentage": r'(n%\s*(to|of)?\s*Net\s*Asset(s)?)|(Rounded\s*%\s*(to|of)?\s*Net\s*Asset(s)?)|%\s*(to|of)?\s*Net\s*Assets|%\s*(of|to)?\s*NAV|%\s*(of|to)?\s*AUM|(Rounded)?\s*(%|Percentage)?\s*(to|of)?\s*Net\s*Assets|%\s*of\s*AUM|%\s*to\s*AUM|%\s*age\s*to\s*NAV',
    }

    for row in sheet.iter_rows(values_only=True):
        row_str = " ".join([str(cell) if cell else "" for cell in row])
        row_str = re.sub(r'\d+(\.\d+)?', '', row_str).strip()
        if not headers_found:
            if re.search(header_pattern, row_str, re.IGNORECASE):
                headers_found = True
                header_row = row
                continue
            else:
                continue

        if headers_found:
            if re.search(stop_pattern, row_str, re.IGNORECASE):
                break

            if re.search(equity_related_pattern, row_str, re.IGNORECASE):
                content_start = True
                continue

            if content_start and re.search(listed_pattern, row_str, re.IGNORECASE):
                continue

            if content_start and any(keyword in row_str for keyword in ["Arbitrage & Special Situations", "Equity", "equity", "Equity Shares"]):
                continue

            if content_start:
                if not any(row):
                    if not first_blank:
                        first_blank = True
                        continue
                    else:
                        break
                data.append(row)

    header_row = [str(col) if col else "" for col in header_row]
    df = pd.DataFrame(data, columns=header_row)

    column_mapping = {}
    for col in df.columns:
        col_str = str(col) if col else ""
        for key, pattern in desired_columns_patterns.items():
            if re.search(pattern, col_str, re.IGNORECASE):
                column_mapping[col] = key
                break

    df.rename(columns=column_mapping, inplace=True)

    df = df[[key for key in desired_columns_patterns.keys() if key in df.columns]]

    sub_total_index = None
    for col in df.columns:
        match_index = df[df[col].astype(str).str.contains(r'\b(?:Sub\s*Total|Subtotal|Total|Unlisted)\b', na=False, case=False)].index
        if not match_index.empty:
            sub_total_index = match_index[0]
            break

    if sub_total_index is not None:
        df = df.iloc[:sub_total_index]

    df = df[df['quantity'] >= 0]
 
    if 'isin' in df.columns or 'quantity' in df.columns:
        df = df.dropna(subset=['isin', 'quantity'], how='all')
    df = df.dropna(subset=['isin'])

    if 'quantity' in df.columns:
        df = df[pd.to_numeric(df['quantity'], errors='coerce').notnull()]
        
    return df

def process_excel_file(file_path, schemeRow, schemeCol, amc_name, percentage_type):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_data = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Processing {sheet_name}.")

        if amc_name == "MIRAE":
            value_in_5th_row = sheet.cell(row=5, column=2).value
            if isinstance(value_in_5th_row, str) and value_in_5th_row.strip().lower().startswith("mirae"):
                scheme_name = value_in_5th_row.strip()
            else:
                scheme_name = sheet.cell(row=6, column=2).value
        else:        
            scheme_name = sheet.cell(row=schemeRow, column=schemeCol).value

        if scheme_name:
            scheme_name = extract_scheme_name(scheme_name, amc_name)
        
        equity = sheet_contains_equity_related(sheet)
        if equity:
            cleaned_data = clean_excel_sheet(sheet)
            if cleaned_data is not None and not cleaned_data.empty:
                if "holding_percentage" in cleaned_data.columns:
            # Replace specific values in the 'holding_percentage' column
                    cleaned_data['holding_percentage'] = cleaned_data['holding_percentage'].replace({'$0.00%': '0.00','00.00%': '0.00', '0.00%': '0.00','$': '0.00','*': '0.00', '@': '0.00', '^': '0.00'})

                    if percentage_type == "dec":
                        # Multiply by 100 and round to 2 decimal places for "dec" type
                        cleaned_data['holding_percentage'] = cleaned_data['holding_percentage'].astype(float) * 100
                        cleaned_data['holding_percentage'] = cleaned_data['holding_percentage'].round(2)
                    # If percentage_type == "%" then no change needed

                if not cleaned_data.empty:
                    cleaned_data['amc_short_name'] = amc_name
                    if amc_name == 'UTI':
                        cleaned_data['mf_short_name'] = os.path.basename(file_path).split('_')[0]
                    else:
                        cleaned_data['mf_short_name'] = sheet_name
                    cleaned_data['fund_name'] = scheme_name
                    cleaned_data['fund_type'] = 'equity'
                    cleaned_data['month_year'] = (datetime.now() - pd.DateOffset(months=1)).strftime('%b-%Y')
                    all_data.append(cleaned_data.reset_index(drop=True))
        else:
            print(f"Skipping sheet {sheet_name} as it does not contain 'Equity & Equity related' or a similar suffix.")

    if all_data:
        combined_df = pd.concat(all_data, ignore_index=True)
        return combined_df
    else:
        return pd.DataFrame()
    
def run_script(script_name, directory_path, output_file_path):
    script_path = os.path.join(r"../CC-DataScience/singleamc", script_name)
    subprocess.run(["python", script_path, directory_path, output_file_path])


def process_amc_directory(directory_path, output_file_path):
    all_amc_data = []
    amc_config = {
        "360ONE": (1, 2, "dec"),
        "ADITYA_BIRLA_CAPITAL": (1, 2, "dec"),
        "AXIS": (1, 2, "dec"),
        "BANDHAN": (3, 2, "dec"),
        "BARODA": (3, 4, "%"),
        "BOI": (1, 2, "dec"),
        "CANARA_ROBECO": (1, 2, "%"),
        "DSP": (1, 2, "dec"),
        "EDELWEISS": (1, 1, "dec"),
        "FRANKLIN_TEMPLETON": (1, 1, "%"),
        "GROWW": (1, 2, "dec"),
        "HDFC": (1, 1, "%"),
        "HELIOS": (3, 4, "%"),
        "HSBC": (2, 1, "dec"),
        "ICICI": (2, 2, "dec"),
        "INVESCO": (3, 2, "%"),
        "ITIAMC": (3, 2, "dec"),
        "JM_FINANCIAL": (2, 1, "%"),
        "KOTAK": (1, 3, "%"),
        "LIC": (1, 2, "dec"),
        "MAHINDRA": (3, 2, "%"),
        "MIRAE": (5, 2, "dec"),
        "MOTILAL": (1, 2, "dec"),
        "NAVI": (3, 1, "%"),
        "NIPPON": (1, 2, "dec"),
        "NJ": (2, 2, "dec"),
        "OLD_BRIDGE_ASSET": (3, 2, "%"),
        "PGIM": (1, 2, "%"),
        "PPFAS": (1, 2, "dec"),
        "QUANT": (2, 3, "%"),
        "SAMCO": (1, 2, "dec"),
        "SBI": (3, 4, "%"),
        "SHRIRAM": (1, 2, "%"),
        "SUNDARAM": (2, 1, "dec"),
        "TATA": (1, 2, "%"),
        "TAURUS": (3, 4, "%"),
        "TRUST": (4, 1, "dec"),
        "UNION": (7, 3, "%"),
        "UTI": (2, 1, "%"),
        "WHITEOAK": (1, 2, "dec"),
        "ZERODHA": (2, 3, "dec"),
        "QUANTUM": (1, 1, "%")
    }

    for amc_name, (row, col, percentage_type) in amc_config.items():
        sub_dir = os.path.join(directory_path, amc_name)
        if not os.path.isdir(sub_dir):
            print(f"Directory {sub_dir} not found, skipping.")
            continue

        # Construct paths for the current AMC
        current_directory_path = os.path.join(directory_path, amc_name)
        current_output_file_path = os.path.join(output_file_path, f"{amc_name}.csv")

        if amc_name in ("PGIM", "TATA", "KOTAK", "SHRIRAM", "SUNDARAM", "FRANKLIN_TEMPLETON", "QUANTUM"):
            script_name = f"{amc_name}.py"  # Construct script name based on amc_name
            run_script(script_name, current_directory_path, current_output_file_path)
        else:
            all_data = []
            for file in os.listdir(sub_dir):
                if file.endswith('.xlsx'):
                    file_path = os.path.join(sub_dir, file)
                    print(f"Processing file: {file} in {sub_dir}")
                    amc_data = process_excel_file(file_path, row, col, amc_name, percentage_type)
                    print(f"Processed data shape: {amc_data.shape if amc_data is not None else 'No data'}")
                    if amc_data is not None and not amc_data.empty:
                        all_data.append(amc_data)

            if all_data:
                combined_df = pd.concat(all_data, ignore_index=True)
                output_file = os.path.join(output_file_path, f"{amc_name}.csv")
                # Save the DataFrame directly to CSV, overwriting if the file exists
                combined_df.to_csv(output_file, index=False)
                print(f"Combined data saved to {output_file}")
                all_amc_data.append(combined_df)
            else:
                print(f"No valid data found for AMC {amc_name}.")

    if all_amc_data:
        combined_all_amc_df = pd.concat(all_amc_data, ignore_index=True)
        # Save the final combined DataFrame, overwriting if the file exists
        combined_all_amc_df.to_csv(output_file_path, index=False)
        print(f"Combined data saved to {output_file_path}")
    else:
        print("No data to save in the combined output file.")

directory_path = MONTHLY_DIRECTORY_PATH
output_file_path = MONTHLY_OUTPUT_PATH

process_amc_directory(directory_path, output_file_path)